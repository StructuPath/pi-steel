#!/usr/bin/env python3
"""Compile a deterministic, draft-only steel RFQ workbook."""

from __future__ import annotations

import argparse
import hashlib
import importlib.util
import json
import os
import re
import sys
from datetime import date
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


SHARED_ROOT = Path(__file__).resolve().parents[2] / "_shared"
if str(SHARED_ROOT) not in sys.path:
    sys.path.insert(0, str(SHARED_ROOT))
from bootstrap import bootstrap_shared  # noqa: E402

bootstrap_shared(__file__)
from pi_steel import RunPublisher, canonical_json_bytes, outcome_exit_code, sha256_bytes  # noqa: E402
from pi_steel.contracts import ESTIMATE_PACKAGE_VERSION, estimate_input_hash  # noqa: E402
from pi_steel.validation import validate_estimate_package  # noqa: E402

RECALC_PATH = Path(__file__).with_name("recalc.py")
RECALC_SPEC = importlib.util.spec_from_file_location("pi_steel_rfq_recalc", RECALC_PATH)
recalc = importlib.util.module_from_spec(RECALC_SPEC)
RECALC_SPEC.loader.exec_module(recalc)


RFQ_COMPILER_VERSION = "1.0.0"
NEST_HANDOFF_VERSION = "1.0.0"
HEADERS = [
    "Item",
    "Category",
    "Description",
    "Size / Designation",
    "Grade",
    "Qty",
    "Stock Length / Size",
    "Est. Purchase Wt (lbs)",
    "Unit Price ($)",
    "Total Price ($)",
    "Availability",
    "Lead Time (days)",
    "Alternate Size",
    "Notes",
]
COLUMN_WIDTHS = {
    "A": 7,
    "B": 13,
    "C": 38,
    "D": 22,
    "E": 12,
    "F": 6,
    "G": 20,
    "H": 20,
    "I": 14,
    "J": 14,
    "K": 14,
    "L": 14,
    "M": 18,
    "N": 30,
}
CATEGORY_ORDER = {
    "W-SHAPES": 10,
    "LONG PRODUCTS": 20,
    "PLATE STOCK": 30,
    "PLATE PARTS": 40,
    "FLAT BAR STOCK": 50,
    "HARDWARE": 60,
    "OTHER": 90,
}


class RfqInputError(ValueError):
    """Raised when an input cannot be mapped without guessing."""


class StageArgumentParser(argparse.ArgumentParser):
    def error(self, message):
        self.print_usage(sys.stderr)
        self.exit(1, f"{self.prog}: error: {message}\n")


def _valid_date(value: Any) -> bool:
    try:
        return date.fromisoformat(str(value)).isoformat() == str(value)
    except ValueError:
        return False


def _fmt(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return f"{value:g}"
    return str(value)


def _category_for(item: dict[str, Any]) -> str:
    dimensions = item.get("dimensions") or {}
    if dimensions.get("category"):
        return str(dimensions["category"]).upper()
    designation = str(item.get("designation") or item.get("specification") or "")
    normalized = designation.upper().replace(" ", "")
    if normalized.startswith("W"):
        return "W-SHAPES"
    if normalized.startswith(("PL", "PLATE")) or item.get("geometry"):
        return "PLATE PARTS"
    if normalized.startswith(("FB", "FLATBAR")):
        return "FLAT BAR STOCK"
    if item.get("intent") == "hardware":
        return "HARDWARE"
    return "OTHER"


def _size_for(item: dict[str, Any]) -> str:
    dimensions = item.get("dimensions") or {}
    if dimensions.get("size"):
        return str(dimensions["size"])
    if item.get("designation"):
        return str(item["designation"])
    if item.get("specification"):
        return str(item["specification"])
    geometry = item.get("geometry") or {}
    if geometry:
        return " x ".join(
            _fmt(geometry.get(field)) for field in ("width", "height", "thickness")
        )
    return ""


def _normalized_item(item: dict[str, Any]) -> dict[str, Any]:
    dimensions = item.get("dimensions") or {}
    geometry = item.get("geometry") or {}
    quantity = item["quantity"]
    weight = dimensions.get("purchase_weight_lbs", item.get("total_weight_lbs"))
    if weight is None and item.get("length_ft") is not None and item.get("unit_weight_plf") is not None:
        weight = quantity * item["length_ft"] * item["unit_weight_plf"]
    stock_length = dimensions.get("stock_length")
    if stock_length is None and item.get("length_ft") is not None:
        stock_length = f"{quantity} pcs x {_fmt(item['length_ft'])} ft"
    return {
        "source_id": item["source_id"],
        "item_id": item["item_id"],
        "intent": item["intent"],
        "item": item.get("mark") or item["source_id"],
        "category": _category_for(item),
        "description": item.get("description") or item.get("mark") or item["source_id"],
        "size": _size_for(item),
        "material": item.get("material", ""),
        "grade": item.get("grade", ""),
        "thickness": dimensions.get("thickness", geometry.get("thickness")),
        "quantity": quantity,
        "stock_length": stock_length or "",
        "purchase_weight_lbs": weight,
        "replaces_item_ids": tuple(dimensions.get("replaces_item_ids", [])),
    }


def normalize_canonical_package(package: dict[str, Any]) -> dict[str, Any]:
    """Map typed canonical scope to deterministic purchase lines."""
    result = validate_estimate_package(package)
    if result.blockers:
        raise RfqInputError(
            "canonical package is blocked: "
            + "; ".join(finding["message"] for finding in result.blockers)
        )
    source_ids = {item["item_id"] for item in package["items"]}
    items_by_id = {item["item_id"]: item for item in package["items"]}
    replacements: set[str] = set()
    for item in package["items"]:
        if item.get("intent") not in {"purchased_stock", "hardware"}:
            continue
        for replaced_id in (item.get("dimensions") or {}).get(
            "replaces_item_ids", []
        ):
            if replaced_id not in source_ids:
                raise RfqInputError(
                    f"purchase relationship references unknown item {replaced_id!r}"
                )
            if items_by_id[replaced_id]["intent"] != "fabricated_part":
                raise RfqInputError(
                    "purchase relationships may replace fabricated_part items only"
                )
            replacements.add(replaced_id)
    included = []
    for item in package["items"]:
        if item["intent"] in {"allowance", "exclusion", "by_others"}:
            continue
        if item["item_id"] in replacements:
            continue
        included.append(_normalized_item(item))
    included.sort(
        key=lambda item: (
            CATEGORY_ORDER.get(item["category"], 80),
            item["material"],
            item["grade"],
            item["thickness"] if item["thickness"] is not None else -1,
            item["size"],
            item["item_id"],
        )
    )
    return {
        "input_kind": "canonical_estimate_package",
        "input_version": ESTIMATE_PACKAGE_VERSION,
        "input_hash": result.input_hash,
        "project_id": package["project"]["project_id"],
        "project_name": package["project"].get("name")
        or package["project"]["project_id"],
        "revision_id": package["project"]["revision"]["revision_id"],
        "currency": package["commercial_basis"]["currency"],
        "items": included,
        "warnings": [finding["message"] for finding in result.warnings],
    }


LEGACY_HEADERS = [
    "Source_ID",
    "Item_ID",
    "Scope",
    "Description",
    "Material",
    "Grade",
    "Thickness",
    "Size",
    "Qty",
    "Currency",
    "Purchase Weight",
]


def normalize_legacy_xlsx(path: str | Path) -> dict[str, Any]:
    """Conservatively map one exact legacy sheet/header contract."""
    workbook = openpyxl.load_workbook(path, data_only=False, read_only=True)
    if "Steel Takeoff" not in workbook.sheetnames:
        raise RfqInputError("legacy workbook requires a 'Steel Takeoff' sheet")
    sheet = workbook["Steel Takeoff"]
    headers = [sheet.cell(1, column).value for column in range(1, len(LEGACY_HEADERS) + 1)]
    if headers != LEGACY_HEADERS:
        raise RfqInputError(
            "legacy adapter requires exact headers: " + ", ".join(LEGACY_HEADERS)
        )
    items = []
    currencies = set()
    for row_number, values in enumerate(
        sheet.iter_rows(min_row=2, max_col=len(LEGACY_HEADERS), values_only=True),
        start=2,
    ):
        row = dict(zip(LEGACY_HEADERS, values))
        if not any(value is not None for value in values):
            continue
        if not row["Source_ID"] or not row["Item_ID"]:
            raise RfqInputError(
                f"legacy row {row_number} requires Source_ID and Item_ID"
            )
        if row["Scope"] not in {"IN SCOPE", "BY OTHERS", "EXCLUDED"}:
            raise RfqInputError(
                f"legacy row {row_number} has ambiguous Scope {row['Scope']!r}"
            )
        currency = str(row["Currency"] or "").strip().upper()
        if not re.fullmatch(r"[A-Z]{3}", currency):
            raise RfqInputError(
                f"legacy row {row_number} requires an explicit three-letter Currency"
            )
        currencies.add(currency)
        try:
            quantity = int(row["Qty"])
        except (TypeError, ValueError):
            raise RfqInputError(
                f"legacy row {row_number} has invalid Qty"
            ) from None
        if quantity != row["Qty"] or quantity < 0:
            raise RfqInputError(
                f"legacy row {row_number} has invalid Qty"
            )
        intent = (
            "by_others"
            if row["Scope"] == "BY OTHERS"
            else ("exclusion" if row["Scope"] == "EXCLUDED" or quantity == 0 else "purchased_stock")
        )
        if intent in {"by_others", "exclusion"}:
            continue
        category = _category_for(
            {
                "intent": intent,
                "specification": row["Size"],
                "dimensions": {},
            }
        )
        if category == "PLATE PARTS":
            category = "PLATE STOCK"
        item = {
            "source_id": str(row["Source_ID"]),
            "item_id": str(row["Item_ID"]),
            "intent": intent,
            "item": str(row["Source_ID"]),
            "category": category,
            "description": str(row["Description"] or ""),
            "size": str(row["Size"] or ""),
            "material": str(row["Material"] or ""),
            "grade": str(row["Grade"] or ""),
            "thickness": row["Thickness"],
            "quantity": quantity,
            "stock_length": "",
            "purchase_weight_lbs": row["Purchase Weight"],
            "replaces_item_ids": (),
        }
        items.append(item)
    if len(currencies) != 1:
        raise RfqInputError("legacy workbook must use one explicit currency")
    items.sort(
        key=lambda item: (
            CATEGORY_ORDER.get(item["category"], 80),
            item["material"],
            item["grade"],
            item["thickness"] if item["thickness"] is not None else -1,
            item["size"],
            item["item_id"],
        )
    )
    input_path = Path(path)
    input_hash = hashlib.sha256(input_path.read_bytes()).hexdigest()
    return {
        "input_kind": "legacy_xlsx",
        "input_version": "exact-headers-1.0.0",
        "input_hash": input_hash,
        "project_id": f"legacy-workbook:{input_hash[:20]}",
        "project_name": input_path.stem,
        "revision_id": "LEGACY-REVISION",
        "currency": next(iter(currencies)),
        "items": items,
        "warnings": [
            "Legacy XLSX mapping used exact headers; confirm typed scope before issue."
        ],
    }


def _user_config_path() -> Path:
    xdg = os.environ.get("XDG_CONFIG_HOME")
    if xdg:
        return Path(xdg) / "pi-steel" / "company-profile.json"
    if sys.platform == "darwin":
        return (
            Path.home()
            / "Library"
            / "Application Support"
            / "pi-steel"
            / "company-profile.json"
        )
    return Path.home() / ".config" / "pi-steel" / "company-profile.json"


def resolve_company_profile(input_path: str | Path) -> tuple[dict[str, Any] | None, str]:
    candidates = []
    explicit = os.environ.get("PI_STEEL_CONFIG")
    if explicit:
        explicit_path = Path(explicit)
        if not explicit_path.is_file():
            raise RfqInputError(
                "PI_STEEL_CONFIG names a profile file that does not exist"
            )
        candidates.append((explicit_path, "environment"))
    candidates.append(
        (Path(input_path).resolve().parent / ".pi-steel" / "company-profile.json", "project")
    )
    candidates.append((_user_config_path(), "user"))
    for path, source in candidates:
        if not path.is_file():
            continue
        try:
            profile = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError) as exc:
            raise RfqInputError(f"company profile could not be read: {exc}") from exc
        profile["_profile_path"] = str(path)
        return profile, source
    return None, "missing"


def validate_company_profile(
    profile: dict[str, Any] | None, profile_path: str | Path | None = None
) -> list[dict[str, str]]:
    findings = []

    def add(code, message):
        findings.append({"code": code, "severity": "error", "message": message})

    if profile is None:
        add("company_profile_missing", "A company profile is required.")
        return findings
    for field in ("company_name", "city_state"):
        if not profile.get(field):
            add("company_identity_incomplete", f"Profile field {field} is required.")
    terms = profile.get("terms_template")
    if not isinstance(terms, dict):
        add("approved_terms_missing", "An approved terms_template is required.")
        return findings
    for field in (
        "template_id",
        "content",
        "content_hash",
        "approver",
        "approval_date",
        "status",
    ):
        if not terms.get(field):
            add("approved_terms_incomplete", f"Terms field {field} is required.")
    content = terms.get("content", "")
    actual_hash = hashlib.sha256(content.encode("utf-8")).hexdigest()
    if terms.get("content_hash") != actual_hash:
        add(
            "terms_content_hash_mismatch",
            "Terms content changed after approval; approval is invalid.",
        )
    if terms.get("status") != "approved":
        add("terms_not_approved", "Terms template status must be approved.")
    if not _valid_date(terms.get("approval_date")):
        add(
            "terms_approval_date_invalid",
            "Terms approval_date must be an explicit YYYY-MM-DD date.",
        )
    return findings


def validate_nest_handoff(value: dict[str, Any] | None) -> list[dict[str, str]]:
    if value is None:
        return []
    findings = []
    if value.get("schema_version") != NEST_HANDOFF_VERSION:
        findings.append(
            {
                "code": "unsupported_nest_handoff_version",
                "severity": "error",
                "message": "Migrate nesting handoff to version 1.0.0.",
            }
        )
    if value.get("source_nest_result_version") != "1.0.0":
        findings.append(
            {
                "code": "unsupported_nest_result_version",
                "severity": "error",
                "message": "Nesting result version must be 1.0.0.",
            }
        )
    if value.get("geometry_readiness") not in {
        "geometry_verified",
        "reference_only",
        "diagnostic",
    }:
        findings.append(
            {
                "code": "invalid_geometry_readiness",
                "severity": "error",
                "message": "Nesting handoff requires explicit geometry_readiness.",
            }
        )
    if not isinstance(value.get("rows"), list):
        findings.append(
            {
                "code": "invalid_nest_rows",
                "severity": "error",
                "message": "Nesting handoff rows must be an array.",
            }
        )
    return findings


def _color(value):
    if value is None:
        return None
    def primitive(attribute):
        result = getattr(value, attribute, None)
        return result if isinstance(result, (str, int, float, bool)) else None
    return {
        "type": primitive("type"),
        "rgb": primitive("rgb"),
        "indexed": primitive("indexed"),
        "theme": primitive("theme"),
        "tint": primitive("tint"),
    }


def workbook_semantic_projection(path: str | Path) -> dict[str, Any]:
    workbook = openpyxl.load_workbook(path, data_only=False)
    sheets = []
    for sheet in workbook.worksheets:
        cells = []
        styles = []
        style_ids = {}
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is None and not cell.has_style:
                    continue
                style = {
                    "number_format": cell.number_format,
                    "font": {
                            "name": cell.font.name,
                            "size": cell.font.sz,
                            "bold": cell.font.b,
                            "italic": cell.font.i,
                            "color": _color(cell.font.color),
                    },
                    "fill": {
                        "type": cell.fill.fill_type,
                        "fg": _color(cell.fill.fgColor),
                    },
                    "alignment": {
                        "horizontal": cell.alignment.horizontal,
                        "vertical": cell.alignment.vertical,
                        "wrap_text": cell.alignment.wrap_text,
                    },
                    "border": {
                        side: getattr(cell.border, side).style
                        for side in ("left", "right", "top", "bottom")
                    },
                }
                style_key = json.dumps(style, sort_keys=True, separators=(",", ":"))
                if style_key not in style_ids:
                    style_ids[style_key] = len(styles)
                    styles.append(style)
                cells.append([cell.coordinate, cell.value, style_ids[style_key]])
        sheets.append(
            {
                "title": sheet.title,
                "state": sheet.sheet_state,
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
                "merges": sorted(str(value) for value in sheet.merged_cells.ranges),
                "freeze_panes": str(sheet.freeze_panes) if sheet.freeze_panes else None,
                "column_widths": {
                    key: dimension.width
                    for key, dimension in sorted(sheet.column_dimensions.items())
                    if dimension.width is not None
                },
                "print_title_rows": sheet.print_title_rows,
                "page_setup": {
                    "orientation": sheet.page_setup.orientation,
                    "fitToWidth": sheet.page_setup.fitToWidth,
                    "fitToHeight": sheet.page_setup.fitToHeight,
                },
                "styles": styles,
                "cells": cells,
            }
        )
    return {"format_version": "1.0.0", "sheets": sheets}


def _safe_filename(project_name: str) -> str:
    stem = re.sub(r"[^A-Za-z0-9._-]+", "_", project_name).strip("._")
    return f"{stem or 'RFQ'}_RFQ_Material_List.xlsx"


def compile_workbook(
    normalized: dict[str, Any],
    profile: dict[str, Any],
    *,
    nest_handoff: dict[str, Any] | None,
    issued_date: str,
    project_location: str,
    output_directory: str | Path,
    profile_source: str,
    bake: bool,
) -> dict[str, Any]:
    profile_findings = validate_company_profile(
        profile, profile.get("_profile_path")
    )
    if profile_findings:
        raise RfqInputError(
            "company profile blocked: "
            + "; ".join(finding["message"] for finding in profile_findings)
        )
    nest_findings = validate_nest_handoff(nest_handoff)
    if nest_findings:
        raise RfqInputError(
            "nest handoff blocked: "
            + "; ".join(finding["message"] for finding in nest_findings)
        )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "RFQ Draft"
    metadata = workbook.create_sheet("RFQ Metadata")
    metadata.sheet_state = "hidden"
    dark_blue = "1F3864"
    medium_blue = "2E75B6"
    gold = "BF8F00"
    yellow = "FFF2CC"
    light_blue = "D6E4F0"
    green = "E2EFDA"
    white = "FFFFFF"
    thin = Side(style="thin", color="A6A6A6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for column, width in COLUMN_WIDTHS.items():
        sheet.column_dimensions[column].width = width
    for row in (1, 2, 3, 7):
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
        cell = sheet.cell(row, 1)
        cell.fill = PatternFill("solid", fgColor=dark_blue)
        cell.font = Font(name="Arial", color=white, bold=row != 3, size={1: 14, 2: 11}.get(row, 10))
        cell.alignment = Alignment(vertical="center")
        for column in range(1, 15):
            sheet.cell(row, column).fill = PatternFill("solid", fgColor=dark_blue)
    sheet["A1"] = "REQUEST FOR QUOTATION — Structural Steel"
    sheet["A2"] = (
        f"{normalized['project_name']} | {profile['company_name']} — "
        f"{profile['city_state']}"
    )
    sheet["A3"] = (
        f"Date Issued: {issued_date} | Response Requested By: _______________ | "
        f"Project Location: {project_location}"
    )
    sheet.merge_cells("A5:B5")
    sheet.merge_cells("C5:E5")
    sheet.merge_cells("F5:G5")
    sheet.merge_cells("H5:J5")
    sheet.merge_cells("A6:B6")
    sheet.merge_cells("C6:E6")
    sheet.merge_cells("F6:G6")
    sheet.merge_cells("H6:J6")
    sheet["A5"], sheet["C5"] = "Company Name", ""
    sheet["F5"], sheet["H5"] = "Contact Name", ""
    sheet["A6"], sheet["C6"] = "Phone / Email", ""
    sheet["F6"], sheet["H6"] = "Quote Valid Until", ""
    for coordinate in ("C5", "H5", "C6", "H6"):
        sheet[coordinate].fill = PatternFill("solid", fgColor=yellow)
    sheet["A7"] = (
        "Instructions: Complete the yellow vendor-response columns. "
        "This workbook is a draft request only and is not a purchase authorization."
    )
    sheet["A7"].font = Font(name="Arial", size=9, italic=True, color="666666")

    for column, header in enumerate(HEADERS, start=1):
        cell = sheet.cell(8, column, header)
        cell.fill = PatternFill(
            "solid", fgColor=gold if column >= 9 else medium_blue
        )
        cell.font = Font(name="Arial", size=10, bold=True, color=white)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    sheet.freeze_panes = "A9"
    sheet.print_title_rows = "8:8"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True

    row = 9
    material_rows = []
    grouped: dict[tuple[Any, ...], list[dict[str, Any]]] = {}
    for item in normalized["items"]:
        key = (
            item["category"],
            item["material"],
            item["grade"],
            item["thickness"],
            item["size"],
        )
        grouped.setdefault(key, []).append(item)
    for key in sorted(
        grouped,
        key=lambda value: (
            CATEGORY_ORDER.get(value[0], 80),
            value[1],
            value[2],
            value[3] if value[3] is not None else -1,
            value[4],
        ),
    ):
        category, material, grade, thickness, size = key
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
        heading = sheet.cell(
            row,
            1,
            f"{category} — {material} — {grade}"
            + (f" — {_fmt(thickness)}" if thickness is not None else ""),
        )
        heading.fill = PatternFill("solid", fgColor=medium_blue)
        heading.font = Font(name="Arial", bold=True, color=white)
        for column in range(1, 15):
            sheet.cell(row, column).fill = PatternFill("solid", fgColor=medium_blue)
        row += 1
        for item in grouped[key]:
            material_rows.append(row)
            values = [
                item["item"],
                item["category"],
                item["description"],
                item["size"],
                item["grade"],
                item["quantity"],
                item["stock_length"],
                item["purchase_weight_lbs"],
            ]
            for column, value in enumerate(values, start=1):
                sheet.cell(row, column, value)
            sheet.cell(row, 10, f'=IF(I{row}="","",F{row}*I{row})')
            for column in range(1, 15):
                cell = sheet.cell(row, column)
                cell.border = border
                cell.font = Font(name="Arial", size=10)
                cell.alignment = Alignment(vertical="top", wrap_text=column in {3, 7, 14})
                cell.fill = PatternFill(
                    "solid",
                    fgColor=yellow
                    if column >= 9
                    else (light_blue if len(material_rows) % 2 else white),
                )
            sheet.cell(row, 8).number_format = "#,##0"
            sheet.cell(row, 9).number_format = f'"{normalized["currency"]}" #,##0.00'
            sheet.cell(row, 10).number_format = f'"{normalized["currency"]}" #,##0.00'
            row += 1

    if not material_rows:
        raise RfqInputError("no in-scope purchase items remain after typed filtering")
    first_material_row = material_rows[0]
    last_material_row = material_rows[-1]
    total_row = row
    sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=7)
    sheet.cell(total_row, 1, "TOTAL PURCHASE WEIGHT / PRICE")
    sheet.cell(total_row, 1).font = Font(name="Arial", bold=True)
    sheet.cell(total_row, 1).alignment = Alignment(horizontal="right")
    sheet.cell(total_row, 8, f"=SUM(H{first_material_row}:H{last_material_row})")
    sheet.cell(total_row, 10, f"=SUM(J{first_material_row}:J{last_material_row})")
    for column in (8, 10):
        sheet.cell(total_row, column).fill = PatternFill("solid", fgColor=green)
        sheet.cell(total_row, column).font = Font(name="Arial", bold=True)
        sheet.cell(total_row, column).border = border
    sheet.cell(total_row, 8).number_format = "#,##0"
    sheet.cell(total_row, 10).number_format = f'"{normalized["currency"]}" #,##0.00'

    nest_header_row = total_row + 2
    sheet.merge_cells(
        start_row=nest_header_row, start_column=1, end_row=nest_header_row, end_column=14
    )
    sheet.cell(
        nest_header_row, 1, "NESTING / REMNANT REFERENCE (For Fabricator Review)"
    )
    sheet.cell(nest_header_row, 1).font = Font(
        name="Arial", bold=True, color=dark_blue
    )
    nest_columns_row = nest_header_row + 1
    nest_headers = [
        "Material",
        "Grade",
        "Thickness",
        "Sheet Size",
        "Nesting Plan",
        "Remnant Notes",
    ]
    for column, header in enumerate(nest_headers, start=1):
        cell = sheet.cell(nest_columns_row, column, header)
        cell.fill = PatternFill("solid", fgColor=medium_blue)
        cell.font = Font(name="Arial", bold=True, color=white)
        cell.border = border
    nest_row = nest_columns_row + 1
    reference_warning = (
        (nest_handoff or {}).get("geometry_readiness") == "reference_only"
    )
    for entry in (nest_handoff or {}).get("rows", []):
        readiness = entry.get(
            "geometry_readiness",
            (nest_handoff or {}).get("geometry_readiness"),
        )
        reference_warning = reference_warning or readiness == "reference_only"
        values = [
            entry.get("material"),
            entry.get("grade"),
            entry.get("thickness"),
            entry.get("sheet_size"),
            entry.get("nesting_plan"),
            entry.get("drop_notes")
            + (" — REFERENCE ONLY" if readiness == "reference_only" else ""),
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(nest_row, column, value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        nest_row += 1

    terms_header_row = nest_row + 1
    sheet.merge_cells(
        start_row=terms_header_row, start_column=1, end_row=terms_header_row, end_column=14
    )
    sheet.cell(terms_header_row, 1, "TERMS & CONDITIONS — APPROVED TEMPLATE")
    sheet.cell(terms_header_row, 1).font = Font(
        name="Arial", bold=True, color=dark_blue
    )
    terms_row = terms_header_row + 1
    sheet.merge_cells(
        start_row=terms_row, start_column=1, end_row=terms_row, end_column=14
    )
    sheet.cell(terms_row, 1, profile["terms_template"]["content"])
    sheet.cell(terms_row, 1).alignment = Alignment(wrap_text=True, vertical="top")

    metadata["A1"], metadata["B1"] = "Field", "Value"
    metadata["A2"], metadata["B2"] = "Document Status", "DRAFT — NOT SENT OR AWARDED"
    metadata["A3"], metadata["B3"] = "Compiler Version", RFQ_COMPILER_VERSION
    metadata["A4"], metadata["B4"] = "Input Hash", normalized["input_hash"]
    metadata["A5"], metadata["B5"] = "Profile Source", profile_source
    metadata["A6"], metadata["B6"] = "Terms Template ID", profile["terms_template"]["template_id"]
    metadata["A7"], metadata["B7"] = "Terms Content Hash", profile["terms_template"]["content_hash"]
    metadata["A8"], metadata["B8"] = "Terms Approver", profile["terms_template"]["approver"]
    metadata["A9"], metadata["B9"] = "Terms Approval Date", profile["terms_template"]["approval_date"]
    metadata["A10"], metadata["B10"] = "Issued Date", issued_date
    metadata["A11"], metadata["B11"] = (
        "Formula Cache",
        "Recalculate on open requested; QA report records cache status",
    )
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"

    logo_status = "not_configured"
    logo = profile.get("logo")
    if logo:
        profile_path = Path(profile.get("_profile_path", "."))
        logo_path = Path(logo)
        if not logo_path.is_absolute():
            logo_path = profile_path.parent / logo_path
        if logo_path.is_file():
            try:
                image = Image(logo_path)
                image.width, image.height = 90, 36
                sheet.add_image(image, "A1")
                logo_status = "embedded"
            except Exception:
                logo_status = "text_fallback"
        else:
            logo_status = "text_fallback"

    output_directory = Path(output_directory)
    output_directory.mkdir(parents=True, exist_ok=True)
    workbook_path = output_directory / _safe_filename(normalized["project_name"])
    workbook.save(workbook_path)
    recalculation_status = recalc.recalculate(workbook_path, bake=bake)
    return {
        "workbook_path": workbook_path,
        "recalculation_status": recalculation_status,
        "logo_status": logo_status,
        "reference_warning": reference_warning,
        "contract": {
            "first_material_row": first_material_row,
            "last_material_row": last_material_row,
            "total_row": total_row,
            "nest_header_row": nest_header_row,
            "terms_header_row": terms_header_row,
        },
    }


def _package_version() -> str:
    path = Path(__file__).resolve().parents[3] / "package.json"
    try:
        return json.loads(path.read_text(encoding="utf-8"))["version"]
    except (OSError, KeyError, json.JSONDecodeError):
        return "unknown"


def _load_normalized(input_path: Path) -> dict[str, Any]:
    if input_path.suffix.lower() == ".json":
        value = json.loads(input_path.read_text(encoding="utf-8"))
        return normalize_canonical_package(value)
    if input_path.suffix.lower() == ".xlsx":
        return normalize_legacy_xlsx(input_path)
    raise RfqInputError("input must be a canonical .json package or exact legacy .xlsx")


def publish_rfq_run(args) -> tuple[dict[str, Any], Path]:
    input_path = Path(args.input)
    input_findings = []
    if not _valid_date(args.issued_date):
        input_findings.append(
            {
                "code": "invalid_issued_date",
                "severity": "error",
                "message": "issued-date must be an explicit YYYY-MM-DD date",
            }
        )
    try:
        normalized = _load_normalized(input_path)
    except (RfqInputError, json.JSONDecodeError) as exc:
        normalized = {
            "input_kind": "invalid",
            "input_version": "unknown",
            "input_hash": hashlib.sha256(input_path.read_bytes()).hexdigest(),
            "project_id": input_path.stem,
            "project_name": input_path.stem,
            "revision_id": "unknown",
            "currency": "USD",
            "items": [],
            "warnings": [],
        }
        input_findings.append(
            {
                "code": "invalid_rfq_input",
                "severity": "error",
                "message": str(exc),
            }
        )
    try:
        profile, profile_source = resolve_company_profile(input_path)
    except RfqInputError as exc:
        profile, profile_source = None, "invalid"
        input_findings.append(
            {
                "code": "invalid_company_profile",
                "severity": "error",
                "message": str(exc),
            }
        )
    profile_findings = validate_company_profile(
        profile, profile.get("_profile_path") if profile else None
    )
    try:
        nest_handoff = (
            json.loads(Path(args.nest).read_text(encoding="utf-8"))
            if args.nest
            else None
        )
        nest_findings = validate_nest_handoff(nest_handoff)
    except (OSError, json.JSONDecodeError) as exc:
        nest_handoff = None
        nest_findings = [
            {
                "code": "invalid_nest_handoff",
                "severity": "error",
                "message": str(exc),
            }
        ]
    findings = input_findings + profile_findings + nest_findings
    blockers = [finding for finding in findings if finding["severity"] == "error"]
    review_reasons = list(normalized["warnings"])
    if any(
        readiness == "reference_only"
        for readiness in [
            (nest_handoff or {}).get("geometry_readiness"),
            *[
                row.get("geometry_readiness")
                for row in (nest_handoff or {}).get("rows", [])
            ],
        ]
    ):
        review_reasons.append("Nesting handoff contains reference-only geometry.")
    if blockers:
        outcome, package_status = "blocked", "draft"
    elif review_reasons:
        outcome, package_status = "review_required", "rfq_draft_review_required"
    else:
        outcome, package_status = "ready", "rfq_ready_for_review"
    profile_hash = (
        sha256_bytes(canonical_json_bytes({k: v for k, v in profile.items() if not k.startswith("_")}))
        if profile
        else sha256_bytes(b"missing-profile")
    )
    configuration_hash = sha256_bytes(
        canonical_json_bytes(
            {
                "compiler_version": RFQ_COMPILER_VERSION,
                "issued_date": args.issued_date,
                "project_location": args.project_location,
                "profile_hash": profile_hash,
                "nest_handoff": nest_handoff,
                "bake_requested": not args.no_bake,
            }
        )
    )
    qa_report = {
        "schema_version": "1.0.0",
        "stage": "steel-rfq",
        "run_outcome": outcome,
        "package_status": package_status,
        "document_status": "DRAFT — NOT SENT OR AWARDED",
        "profile_source": profile_source,
        "findings": findings,
        "warnings": review_reasons,
    }
    with RunPublisher(
        args.out,
        stage="steel-rfq",
        run_outcome=outcome,
        package_status=package_status,
        input_hash=normalized["input_hash"],
        configuration_hash=configuration_hash,
        schema_versions={
            "run_manifest": "1.0.0",
            "estimate_package": normalized["input_version"],
            "rfq_workbook": RFQ_COMPILER_VERSION,
            "rfq_nesting": NEST_HANDOFF_VERSION,
        },
        tool_versions={
            "pi_steel": _package_version(),
            "rfq_compiler": RFQ_COMPILER_VERSION,
        },
        explicit_dates={"issued_date": args.issued_date},
        warnings=review_reasons
        + [finding["message"] for finding in findings],
        approximations=[],
        run_id=args.run_id,
    ) as publisher:
        compile_result = None
        if not blockers:
            compile_result = compile_workbook(
                normalized,
                profile,
                nest_handoff=nest_handoff,
                issued_date=args.issued_date,
                project_location=args.project_location,
                output_directory=publisher.staging_path,
                profile_source=profile_source,
                bake=not args.no_bake,
            )
            qa_report["recalculation_status"] = compile_result["recalculation_status"]
            qa_report["logo_status"] = compile_result["logo_status"]
            workbook_name = compile_result["workbook_path"].name
            publisher.register_artifact(
                workbook_name,
                readiness="draft",
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            publisher.write_json(
                "workbook-semantic.json",
                workbook_semantic_projection(compile_result["workbook_path"]),
                readiness="diagnostic",
            )
        publisher.write_qa_report(qa_report)
        final_path = publisher.publish()
    return qa_report, final_path


def main(argv=None) -> int:
    parser = StageArgumentParser(description=__doc__)
    parser.add_argument("--input", required=True)
    parser.add_argument("--nest")
    parser.add_argument("--out", default="out")
    parser.add_argument("--issued-date", required=True)
    parser.add_argument("--project-location", default="")
    parser.add_argument("--no-bake", action="store_true")
    parser.add_argument("--run-id", help=argparse.SUPPRESS)
    args = parser.parse_args(argv)
    try:
        qa_report, final_path = publish_rfq_run(args)
    except (OSError, json.JSONDecodeError, RfqInputError) as exc:
        print(f"RFQ generation failed: {exc}", file=sys.stderr)
        return 1
    print(f"Published {qa_report['run_outcome']} draft RFQ run: {final_path}")
    return outcome_exit_code(qa_report["run_outcome"])


if __name__ == "__main__":
    raise SystemExit(main())
