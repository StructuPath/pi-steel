import importlib.util
import hashlib
import json
import sys
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
SCRIPT = ROOT / "skills" / "steel-rfq" / "scripts" / "generate-rfq.py"
SPEC = importlib.util.spec_from_file_location("pi_steel_rfq_workbook", SCRIPT)
rfq = importlib.util.module_from_spec(SPEC)
sys.modules[SPEC.name] = rfq
SPEC.loader.exec_module(rfq)
FIXTURES = ROOT / "tests" / "fixtures" / "rfq"
GOLDEN = ROOT / "tests" / "golden" / "rfq" / "semantic-workbook.json"


def load(name):
    return json.loads((FIXTURES / name).read_text())


def build(tmp_path, monkeypatch):
    package = load("estimate-package.json")
    normalized = rfq.normalize_canonical_package(package)
    profile = load("synthetic-profile.json")
    nest = load("nest-handoff.json")
    monkeypatch.setattr(rfq.recalc, "libreoffice_bake", lambda path: False)
    return rfq.compile_workbook(
        normalized,
        profile,
        nest_handoff=nest,
        issued_date="2026-07-28",
        project_location="Example City, ST",
        output_directory=tmp_path,
        profile_source="environment",
        bake=True,
    )


def test_workbook_structure_formulas_styles_and_draft_metadata(tmp_path, monkeypatch):
    result = build(tmp_path, monkeypatch)
    workbook = openpyxl.load_workbook(result["workbook_path"], data_only=False)
    sheet = workbook["RFQ Draft"]
    assert sheet.merged_cells.ranges
    assert {"A1:N1", "A2:N2", "A3:N3", "A7:N7"} <= {
        str(value) for value in sheet.merged_cells.ranges
    }
    assert sheet.freeze_panes == "A9"
    assert sheet.print_title_rows == "$8:$8"
    assert sheet.page_setup.orientation == "landscape"
    assert sheet.page_setup.fitToWidth == 1
    assert sheet.column_dimensions["C"].width == 38
    assert sheet["I8"].fill.fgColor.rgb.endswith("BF8F00")
    assert sheet["I10"].fill.fgColor.rgb.endswith("FFF2CC")
    total_row = result["contract"]["total_row"]
    first_row = result["contract"]["first_material_row"]
    last_row = result["contract"]["last_material_row"]
    assert sheet[f"H{total_row}"].value == f"=SUM(H{first_row}:H{last_row})"
    assert sheet[f"J{total_row}"].value == f"=SUM(J{first_row}:J{last_row})"
    assert workbook["RFQ Metadata"]["B2"].value == "DRAFT — NOT SENT OR AWARDED"
    assert result["recalculation_status"] == "deferred_recalculate_on_open"
    assert result["logo_status"] == "text_fallback"


def test_nesting_rows_stay_separate_and_reference_warning_is_visible(
    tmp_path, monkeypatch
):
    result = build(tmp_path, monkeypatch)
    workbook = openpyxl.load_workbook(result["workbook_path"], data_only=False)
    sheet = workbook["RFQ Draft"]
    values = [
        tuple(sheet.cell(row=row, column=column).value for column in range(1, 7))
        for row in range(1, sheet.max_row + 1)
    ]
    nest_rows = [row for row in values if row[0] in {"carbon_steel"}]
    assert len(nest_rows) == 2
    assert {(row[1], row[2], row[3]) for row in nest_rows} == {
        ("A36", 0.5, "48x24"),
        ("A572", 0.375, "60x30"),
    }
    assert any(
        "REFERENCE ONLY" in str(cell).upper()
        for row in values
        for cell in row
        if cell
    )


def test_semantic_projection_matches_golden(tmp_path, monkeypatch):
    result = build(tmp_path, monkeypatch)
    actual = rfq.workbook_semantic_projection(result["workbook_path"])
    expected = json.loads(GOLDEN.read_text())
    canonical = json.dumps(
        actual, sort_keys=True, separators=(",", ":"), ensure_ascii=False
    ).encode()
    assert hashlib.sha256(canonical).hexdigest() == expected["semantic_sha256"]
    assert [
        {
            "title": sheet["title"],
            "max_row": sheet["max_row"],
            "max_column": sheet["max_column"],
            "cell_count": len(sheet["cells"]),
            "style_count": len(sheet["styles"]),
        }
        for sheet in actual["sheets"]
    ] == expected["sheets"]


def test_formula_like_input_is_literal_but_compiler_formulas_remain_formulas(
    tmp_path, monkeypatch
):
    package = load("estimate-package.json")
    package["project"]["name"] = "=1+1"
    package["items"][0]["description"] = "@SUM(1,1)"
    normalized = rfq.normalize_canonical_package(package)
    profile = load("synthetic-profile.json")
    profile["terms_template"]["content"] = "+DANGEROUS"
    profile["terms_template"]["content_hash"] = hashlib.sha256(
        profile["terms_template"]["content"].encode()
    ).hexdigest()
    monkeypatch.setattr(rfq.recalc, "libreoffice_bake", lambda path: False)

    result = rfq.compile_workbook(
        normalized,
        profile,
        nest_handoff=None,
        issued_date="2026-07-28",
        project_location="-1+1",
        output_directory=tmp_path,
        profile_source="environment",
        bake=True,
    )

    workbook = openpyxl.load_workbook(result["workbook_path"], data_only=False)
    sheet = workbook["RFQ Draft"]
    untrusted = [
        cell
        for row in sheet.iter_rows()
        for cell in row
        if isinstance(cell.value, str)
        and any(token in cell.value for token in ("=1+1", "@SUM", "+DANGEROUS"))
    ]
    assert untrusted
    assert all(cell.data_type == "s" and cell.value.startswith("'") for cell in untrusted)
    total_row = result["contract"]["total_row"]
    assert sheet[f"H{total_row}"].data_type == "f"
    assert sheet[f"J{total_row}"].data_type == "f"


def test_canonical_review_findings_and_assumptions_are_visible(
    tmp_path, monkeypatch
):
    package = load("estimate-package.json")
    package["assumptions"] = [
        {
            "assumption_id": "SYNTHETIC-ASSUMPTION-1",
            "text": "Synthetic basis requires confirmation.",
            "status": "unresolved",
        }
    ]
    normalized = rfq.normalize_canonical_package(package)
    profile = load("synthetic-profile.json")
    monkeypatch.setattr(rfq.recalc, "libreoffice_bake", lambda path: False)

    result = rfq.compile_workbook(
        normalized,
        profile,
        nest_handoff=None,
        issued_date="2026-07-28",
        project_location="Example City, ST",
        output_directory=tmp_path,
        profile_source="environment",
        bake=True,
    )

    workbook = openpyxl.load_workbook(result["workbook_path"], data_only=False)
    visible_text = " ".join(
        str(cell.value)
        for row in workbook["RFQ Draft"].iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert "REVIEW NOTES / ASSUMPTIONS" in visible_text
    assert "SYNTHETIC-ASSUMPTION-1" in visible_text
    assert "unresolved" in visible_text
