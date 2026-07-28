import json
import importlib.util
import os
import subprocess
import sys
from copy import deepcopy
from pathlib import Path

import openpyxl
import pytest


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "skills" / "_shared"))
from pi_steel.contracts import estimate_input_hash

SCRIPT = (
    ROOT
    / "skills"
    / "steel-estimate"
    / "scripts"
    / "build-estimate-package.py"
)
FIXTURES = ROOT / "tests" / "fixtures" / "pipeline"
GOLDEN = ROOT / "tests" / "golden" / "pipeline" / "ready-artifacts.json"
SPEC = importlib.util.spec_from_file_location("pi_steel_estimate_pipeline", SCRIPT)
estimate_pipeline = importlib.util.module_from_spec(SPEC)
sys.modules[SPEC.name] = estimate_pipeline
SPEC.loader.exec_module(estimate_pipeline)


def load_package():
    return json.loads((FIXTURES / "synthetic-estimate.json").read_text())


def run_pipeline(tmp_path, package, run_id, *, profile=True):
    input_path = tmp_path / f"{run_id}-input.json"
    input_path.write_text(json.dumps(package), encoding="utf-8")
    output = tmp_path / "published"
    environment = os.environ.copy()
    environment.pop("PYTHONPATH", None)
    environment["XDG_CONFIG_HOME"] = str(tmp_path / "empty-config")
    if profile:
        environment["PI_STEEL_CONFIG"] = str(
            FIXTURES / "synthetic-profile.json"
        )
    else:
        environment.pop("PI_STEEL_CONFIG", None)
    completed = subprocess.run(
        [
            sys.executable,
            SCRIPT,
            "--input",
            input_path,
            "--out",
            output,
            "--prepared-date",
            "2026-07-28",
            "--issued-date",
            "2026-07-29",
            "--project-location",
            "Example City, ST",
            "--run-id",
            run_id,
            "--no-render",
            "--no-bake",
        ],
        cwd=tmp_path,
        env=environment,
        text=True,
        capture_output=True,
    )
    pointer = json.loads((output / "latest-run.json").read_text())
    return completed, output / pointer["run_directory"]


def load_json(path):
    return json.loads(Path(path).read_text())


def test_ready_pipeline_matches_artifact_contract_and_preserves_scope(tmp_path):
    completed, run_path = run_pipeline(
        tmp_path, load_package(), "SYNTHETIC-PIPELINE-READY"
    )
    assert completed.returncode == 0, completed.stdout + completed.stderr
    manifest = load_json(run_path / "run-manifest.json")
    golden = load_json(GOLDEN)
    assert manifest["run_outcome"] == golden["run_outcome"]
    assert manifest["package_status"] == golden["package_status"]
    assert [artifact["path"] for artifact in manifest["artifacts"]] == golden[
        "artifact_paths"
    ]
    normalized = load_json(run_path / "estimate-package.json")
    assert [item["item_id"] for item in normalized["items"]][:2] == [
        "item:synthetic-pipeline-a1",
        "item:synthetic-pipeline-hss1",
    ]
    bom = load_json(run_path / "normalized-bom.json")
    assert {row["intent"] for row in bom["items"]} >= {
        "fabricated_part",
        "purchased_stock",
        "allowance",
        "exclusion",
    }
    nest = load_json(run_path / "nest-result.json")
    handoff = load_json(run_path / "rfq-nesting.json")
    assert nest["estimate_input_hash"] == manifest["input_hash"]
    assert nest["normalized_input_hash"] != nest["estimate_input_hash"]
    assert handoff["project_id"] == "SYNTHETIC-PIPELINE-001"
    assert handoff["revision_id"] == "SYNTHETIC-REV-A"
    assert handoff["estimate_input_hash"] == manifest["input_hash"]
    assert {
        (group["grade"], group["thickness"]) for group in nest["groups"]
    } == {("A36", 0.5), ("A572", 0.375)}
    workbook = openpyxl.load_workbook(next(run_path.glob("*.xlsx")))
    assert workbook["RFQ Metadata"]["B2"].value == "DRAFT — NOT SENT OR AWARDED"


def test_same_inputs_and_dates_have_stable_semantics_across_isolated_runs(tmp_path):
    first, first_path = run_pipeline(
        tmp_path, load_package(), "SYNTHETIC-PIPELINE-RUN-A"
    )
    second, second_path = run_pipeline(
        tmp_path, load_package(), "SYNTHETIC-PIPELINE-RUN-B"
    )
    assert first.returncode == second.returncode == 0
    for name in (
        "estimate-package.json",
        "normalized-bom.json",
        "nest-result.json",
        "rfq-nesting.json",
        "workbook-semantic.json",
    ):
        assert (first_path / name).read_bytes() == (second_path / name).read_bytes()
    first_manifest = load_json(first_path / "run-manifest.json")
    second_manifest = load_json(second_path / "run-manifest.json")
    assert first_manifest["semantic_hash"] == second_manifest["semantic_hash"]
    assert first_path != second_path


def test_revision_or_quantity_change_updates_input_and_downstream_hashes(tmp_path):
    base = load_package()
    _, first_path = run_pipeline(tmp_path, base, "SYNTHETIC-PIPELINE-BASE")
    changed = deepcopy(base)
    changed["project"]["revision"]["revision_id"] = "SYNTHETIC-REV-B"
    changed["items"][2]["quantity"] = 3
    _, changed_path = run_pipeline(
        tmp_path, changed, "SYNTHETIC-PIPELINE-CHANGED"
    )
    first_manifest = load_json(first_path / "run-manifest.json")
    changed_manifest = load_json(changed_path / "run-manifest.json")
    assert first_manifest["input_hash"] != changed_manifest["input_hash"]
    assert first_manifest["semantic_hash"] != changed_manifest["semantic_hash"]
    first_artifacts = {
        artifact["path"]: artifact["sha256"]
        for artifact in first_manifest["artifacts"]
    }
    changed_artifacts = {
        artifact["path"]: artifact["sha256"]
        for artifact in changed_manifest["artifacts"]
    }
    assert first_artifacts["nest-result.json"] != changed_artifacts["nest-result.json"]


@pytest.mark.parametrize("blocker", ["validation", "unplaced", "profile"])
def test_blockers_preserve_diagnostics_but_never_publish_workbook(tmp_path, blocker):
    package = load_package()
    profile = True
    if blocker == "validation":
        package["items"][0]["quantity"] = 0
    elif blocker == "unplaced":
        package["items"][2]["quantity"] = 100
    else:
        profile = False
    completed, run_path = run_pipeline(
        tmp_path,
        package,
        f"SYNTHETIC-PIPELINE-BLOCKED-{blocker.upper()}",
        profile=profile,
    )
    assert completed.returncode == 3, completed.stdout + completed.stderr
    assert not list(run_path.glob("*.xlsx"))
    manifest = load_json(run_path / "run-manifest.json")
    assert manifest["run_outcome"] == "blocked"
    assert (run_path / "qa-report.json").exists()
    if blocker == "unplaced":
        assert (run_path / "nest-result.json").exists()
        assert load_json(run_path / "nest-result.json")["unplaced"]
    if blocker == "validation":
        assert not (run_path / "nest-result.json").exists()


def test_complete_irregular_nest_yields_review_required_draft(tmp_path):
    package = load_package()
    package["items"][2]["geometry"].update(shape="irregular", area=30)
    completed, run_path = run_pipeline(
        tmp_path, package, "SYNTHETIC-PIPELINE-REFERENCE"
    )
    assert completed.returncode == 2, completed.stdout + completed.stderr
    manifest = load_json(run_path / "run-manifest.json")
    assert manifest["run_outcome"] == "review_required"
    assert manifest["package_status"] == "rfq_draft_review_required"
    qa = load_json(run_path / "qa-report.json")
    assert any(
        approximation["code"] == "BOUNDING_BOX_NESTING"
        for approximation in qa["approximations"]
    )
    workbook = openpyxl.load_workbook(next(run_path.glob("*.xlsx")))
    assert any(
        "REFERENCE ONLY" in str(cell.value).upper()
        for row in workbook["RFQ Draft"].iter_rows()
        for cell in row
        if cell.value
    )


def test_schema_blocker_with_missing_identity_still_publishes_diagnostics(tmp_path):
    package = load_package()
    del package["items"][0]["item_id"]

    completed, run_path = run_pipeline(
        tmp_path, package, "SYNTHETIC-PIPELINE-MISSING-IDENTITY"
    )

    assert completed.returncode == 3, completed.stdout + completed.stderr
    assert not list(run_path.glob("*.xlsx"))
    qa = load_json(run_path / "qa-report.json")
    assert qa["run_outcome"] == "blocked"
    assert any(finding["code"] == "schema_validation" for finding in qa["findings"])
    assert load_json(run_path / "normalized-bom.json")["totals"]["status"] == (
        "blocked_invalid_input"
    )


def test_plate_source_weight_is_not_double_counted_with_nested_weight(tmp_path):
    package = load_package()
    plate = next(item for item in package["items"] if item.get("geometry"))
    plate["total_weight_lbs"] = 999

    completed, run_path = run_pipeline(
        tmp_path, package, "SYNTHETIC-PIPELINE-PLATE-WEIGHT"
    )

    assert completed.returncode == 0, completed.stdout + completed.stderr
    bom = load_json(run_path / "normalized-bom.json")
    assert bom["totals"]["known_member_weight_lbs"] == 899
    assert bom["totals"]["fabricated_weight_lbs"] == (
        bom["totals"]["known_member_weight_lbs"]
        + bom["totals"]["nested_plate_weight_lbs"]
    )


def test_blocked_run_replaces_latest_pointer_without_reusing_ready_artifacts(
    tmp_path,
):
    ready, ready_path = run_pipeline(
        tmp_path, load_package(), "SYNTHETIC-PIPELINE-READY-FIRST"
    )
    assert ready.returncode == 0
    assert list(ready_path.glob("*.xlsx"))

    blocked_package = load_package()
    blocked_package["stock"][0]["quantity"] = 0
    blocked, blocked_path = run_pipeline(
        tmp_path, blocked_package, "SYNTHETIC-PIPELINE-BLOCKED-LATEST"
    )

    assert blocked.returncode == 3
    assert blocked_path != ready_path
    assert not list(blocked_path.glob("*.xlsx"))
    pointer = load_json(tmp_path / "published" / "latest-run.json")
    assert tmp_path / "published" / pointer["run_directory"] == blocked_path


def test_confirmed_on_hand_stock_is_consumed_without_duplicate_rfq_demand(tmp_path):
    package = load_package()
    purchase_ids = {
        item["item_id"]
        for item in package["items"]
        if item["intent"] == "purchased_stock"
    }
    for stock in package["stock"]:
        stock.update(
            stock_kind="on_hand",
            measured_at="2026-07-28",
            source="SYNTHETIC-INVENTORY-COUNT",
        )
    confirmation_hash = estimate_input_hash(package)
    for stock in package["stock"]:
        stock["reviewer_confirmation"] = {
            "actor": "Synthetic Reviewer",
            "timestamp": "2026-07-28T12:00:00Z",
            "estimate_hash": confirmation_hash,
        }

    completed, run_path = run_pipeline(
        tmp_path, package, "SYNTHETIC-PIPELINE-ON-HAND"
    )

    assert completed.returncode == 0, completed.stdout + completed.stderr
    consumption = load_json(run_path / "inventory-consumption.json")
    assert {row["purchase_item_id"] for row in consumption} == purchase_ids
    assert all(row["remaining_purchase_quantity"] == 0 for row in consumption)
    semantic = load_json(run_path / "workbook-semantic.json")
    workbook_text = json.dumps(semantic)
    assert not any(item_id in workbook_text for item_id in purchase_ids)


def test_inventory_consumption_uses_explicit_links_and_never_removes_twice():
    package = {
        "items": [
            {
                "intent": "purchased_stock",
                "item_id": "item:synthetic-stock-a",
                "quantity": 1,
                "dimensions": {"inventory_id": "SYNTHETIC-INV-A"},
            },
            {
                "intent": "purchased_stock",
                "item_id": "item:synthetic-stock-b",
                "quantity": 1,
                "dimensions": {"inventory_id": "SYNTHETIC-INV-B"},
            },
            {
                "intent": "purchased_stock",
                "item_id": "item:synthetic-unlinked",
                "quantity": 1,
                "dimensions": {},
            },
        ]
    }
    normalized = {
        "items": [
            {
                "item_id": item["item_id"],
                "quantity": item["quantity"],
                "purchase_weight_lbs": 10,
            }
            for item in package["items"]
        ]
    }
    nest_result = {
        "plate_reports": [
            {"stock_id": "SYNTHETIC-INV-A"},
            {"stock_id": "SYNTHETIC-INV-B"},
        ]
    }

    lineage = estimate_pipeline.apply_inventory_consumption(
        package,
        normalized,
        nest_result,
        {"SYNTHETIC-INV-A", "SYNTHETIC-INV-B"},
    )

    assert {row["purchase_item_id"] for row in lineage} == {
        "item:synthetic-stock-a",
        "item:synthetic-stock-b",
    }
    assert [item["item_id"] for item in normalized["items"]] == [
        "item:synthetic-unlinked"
    ]


def test_wrong_container_type_still_publishes_blocked_diagnostics(tmp_path):
    package = load_package()
    package["project"] = []

    completed, run_path = run_pipeline(
        tmp_path, package, "SYNTHETIC-PIPELINE-WRONG-CONTAINER"
    )

    assert completed.returncode == 3, completed.stdout + completed.stderr
    assert not list(run_path.glob("*.xlsx"))
    qa = load_json(run_path / "qa-report.json")
    assert qa["run_outcome"] == "blocked"
    assert qa["project_id"] is None
    assert any(finding["path"] == "$.project" for finding in qa["findings"])


def test_no_vendor_supply_items_publish_blocked_diagnostics(tmp_path):
    package = load_package()
    package["items"] = [
        item for item in package["items"] if item["intent"] == "exclusion"
    ]
    package["stock"] = []

    completed, run_path = run_pipeline(
        tmp_path, package, "SYNTHETIC-PIPELINE-NO-SUPPLY"
    )

    assert completed.returncode == 3, completed.stdout + completed.stderr
    assert not list(run_path.glob("*.xlsx"))
    qa = load_json(run_path / "qa-report.json")
    assert any(finding["code"] == "rfq_input_blocked" for finding in qa["findings"])
